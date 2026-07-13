"""
Turn any of the supported inputs into a clean list of Apple Podcast URLs / ids:
  - free text (newline and/or comma separated)
  - an uploaded CSV file
  - an uploaded Excel (.xlsx) file

For files we scan every cell and keep anything that looks like an Apple
Podcasts URL or a bare numeric id, so the user does not have to name a column.
"""

import io
import re

_APPLE_RE = re.compile(r"https?://podcasts\.apple\.com/\S+", re.I)
_ID_RE = re.compile(r"\bid\d{5,}\b|^\d{5,}$")


def from_text(text):
    """Split free text on newlines and commas into candidate URLs/ids."""
    if not text:
        return []
    parts = re.split(r"[\n,]+", text)
    return _dedupe([p.strip() for p in parts if p.strip()])


def from_csv(file_storage):
    """Extract Apple URLs / ids from any cell of an uploaded CSV file."""
    raw = file_storage.read()
    text = raw.decode("utf-8", errors="ignore")
    return _scan_text_blob(text)


def from_excel(file_storage):
    """Extract Apple URLs / ids from any cell of an uploaded .xlsx file."""
    from openpyxl import load_workbook

    data = io.BytesIO(file_storage.read())
    wb = load_workbook(data, read_only=True, data_only=True)
    found = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                found.extend(_match_cell(str(cell)))
    wb.close()
    return _dedupe(found)


def _scan_text_blob(text):
    found = []
    for line in text.splitlines():
        for cell in line.split(","):
            found.extend(_match_cell(cell))
    return _dedupe(found)


def _match_cell(cell):
    cell = cell.strip().strip('"').strip("'")
    if not cell:
        return []
    urls = _APPLE_RE.findall(cell)
    if urls:
        return urls
    if _ID_RE.search(cell):
        return [cell]
    return []


def _dedupe(items):
    seen = set()
    out = []
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out
